"""
办公效率工具箱 (office-helper)

提供10个常用办公自动化工具，包括文件批量重命名、PDF合并与拆分、Excel/CSV互转、
会议纪要生成、排班表生成、文件自动分类、批量文本替换和密码生成。

主要功能:
    - batch_rename_files: 批量重命名文件
    - merge_pdf_files: 合并多个PDF文件
    - split_pdf_file: 拆分PDF文件
    - excel_to_csv: Excel文件转CSV
    - csv_to_excel: CSV转Excel文件
    - generate_meeting_minutes: 生成会议纪要（Markdown格式）
    - create_work_schedule: 生成排班表
    - file_organizer: 文件自动分类整理
    - text_replacer: 批量文本替换
    - password_generator: 密码生成器

依赖:
    - openpyxl: Excel文件读写
    - PyPDF2: PDF文件操作
    - python-docx: Word文档操作
"""

import os
import re
import random
import string
import csv
import shutil
from datetime import datetime, timedelta


# =============================================================================
# 1. 批量重命名文件
# =============================================================================
def batch_rename_files(directory, pattern, new_pattern):
    """
    批量重命名指定目录下的文件。

    使用正则表达式匹配文件名中的模式，并将其替换为新模式。
    例如将 "report_001.txt" 中的 "report" 替换为 "doc"。

    Args:
        directory (str): 目标目录路径。
        pattern (str): 要匹配的正则表达式模式（旧文件名中的模式）。
        new_pattern (str): 替换后的新模式。

    Returns:
        dict: 包含以下键的字典:
            - "total": 匹配到的文件总数
            - "renamed": 成功重命名的文件列表（每项为 {"old": 旧路径, "new": 新路径}）
            - "skipped": 跳过的文件列表（文件名不匹配的文件）
            - "errors": 错误信息列表

    Example:
        >>> batch_rename_files("/data", r"img_(\d+)", r"photo_\1")
        {'total': 5, 'renamed': [...], 'skipped': [], 'errors': []}
    """
    result = {"total": 0, "renamed": [], "skipped": [], "errors": []}

    if not os.path.isdir(directory):
        result["errors"].append(f"目录不存在: {directory}")
        return result

    for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)
        if not os.path.isfile(filepath):
            continue

        result["total"] += 1

        if re.search(pattern, filename):
            new_filename = re.sub(pattern, new_pattern, filename)
            new_filepath = os.path.join(directory, new_filename)
            try:
                os.rename(filepath, new_filepath)
                result["renamed"].append({"old": filepath, "new": new_filepath})
            except OSError as e:
                result["errors"].append(f"重命名失败 {filename}: {e}")
        else:
            result["skipped"].append(filename)

    return result


# =============================================================================
# 2. 合并PDF文件
# =============================================================================
def merge_pdf_files(pdf_list, output_path):
    """
    将多个PDF文件合并为一个PDF文件。

    Args:
        pdf_list (list): PDF文件路径列表，按顺序合并。
        output_path (str): 合并后的输出PDF文件路径。

    Returns:
        dict: 包含以下键的字典:
            - "success": 是否成功（布尔值）
            - "output": 输出文件路径
            - "merged_count": 合并的文件数量
            - "errors": 错误信息列表

    Example:
        >>> merge_pdf_files(["a.pdf", "b.pdf"], "merged.pdf")
        {'success': True, 'output': 'merged.pdf', 'merged_count': 2, 'errors': []}
    """
    from PyPDF2 import PdfWriter, PdfReader

    result = {"success": False, "output": output_path, "merged_count": 0, "errors": []}

    writer = PdfWriter()

    for pdf_path in pdf_list:
        if not os.path.isfile(pdf_path):
            result["errors"].append(f"文件不存在: {pdf_path}")
            continue
        try:
            reader = PdfReader(pdf_path)
            for page in reader.pages:
                writer.add_page(page)
            result["merged_count"] += 1
        except Exception as e:
            result["errors"].append(f"读取PDF失败 {pdf_path}: {e}")

    if result["merged_count"] == 0:
        result["errors"].append("没有成功读取任何PDF文件")
        return result

    try:
        with open(output_path, "wb") as f:
            writer.write(f)
        result["success"] = True
    except Exception as e:
        result["errors"].append(f"写入输出PDF失败: {e}")

    return result


# =============================================================================
# 3. 拆分PDF文件
# =============================================================================
def split_pdf_file(pdf_path, page_ranges, output_dir):
    """
    将PDF文件按指定页码范围拆分为多个PDF文件。

    Args:
        pdf_path (str): 源PDF文件路径。
        page_ranges (list): 页码范围列表，每项为 (start, end) 元组，页码从1开始。
            例如 [(1, 3), (4, 6)] 表示将1-3页和4-6页分别拆分。
        output_dir (str): 输出目录路径。

    Returns:
        dict: 包含以下键的字典:
            - "success": 是否成功
            - "output_files": 生成的文件路径列表
            - "total_pages": 源PDF总页数
            - "errors": 错误信息列表

    Example:
        >>> split_pdf_file("report.pdf", [(1, 2), (3, 5)], "./output")
        {'success': True, 'output_files': ['./output/split_1_2.pdf', ...], ...}
    """
    from PyPDF2 import PdfReader, PdfWriter

    result = {"success": False, "output_files": [], "total_pages": 0, "errors": []}

    if not os.path.isfile(pdf_path):
        result["errors"].append(f"文件不存在: {pdf_path}")
        return result

    os.makedirs(output_dir, exist_ok=True)

    try:
        reader = PdfReader(pdf_path)
        result["total_pages"] = len(reader.pages)
    except Exception as e:
        result["errors"].append(f"读取PDF失败: {e}")
        return result

    for start, end in page_ranges:
        writer = PdfWriter()
        # 页码从1开始，转换为0-based索引
        for page_num in range(start - 1, min(end, result["total_pages"])):
            writer.add_page(reader.pages[page_num])

        output_filename = f"split_{start}_{end}.pdf"
        output_filepath = os.path.join(output_dir, output_filename)
        try:
            with open(output_filepath, "wb") as f:
                writer.write(f)
            result["output_files"].append(output_filepath)
        except Exception as e:
            result["errors"].append(f"写入拆分PDF失败 {output_filename}: {e}")

    result["success"] = len(result["output_files"]) > 0
    return result


# =============================================================================
# 4. Excel转CSV
# =============================================================================
def excel_to_csv(excel_path, output_dir):
    """
    将Excel文件中的所有工作表转换为CSV文件。

    每个工作表会生成一个独立的CSV文件，文件名为 "原文件名_工作表名.csv"。

    Args:
        excel_path (str): Excel文件路径（.xlsx 或 .xls）。
        output_dir (str): 输出CSV文件的目录路径。

    Returns:
        dict: 包含以下键的字典:
            - "success": 是否成功
            - "output_files": 生成的CSV文件路径列表
            - "sheet_count": 转换的工作表数量
            - "errors": 错误信息列表

    Example:
        >>> excel_to_csv("data.xlsx", "./csv_output")
        {'success': True, 'output_files': ['data_Sheet1.csv'], 'sheet_count': 1, 'errors': []}
    """
    from openpyxl import load_workbook

    result = {"success": False, "output_files": [], "sheet_count": 0, "errors": []}

    if not os.path.isfile(excel_path):
        result["errors"].append(f"文件不存在: {excel_path}")
        return result

    os.makedirs(output_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(excel_path))[0]

    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        result["errors"].append(f"读取Excel失败: {e}")
        return result

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        csv_filename = f"{base_name}_{sheet_name}.csv"
        csv_filepath = os.path.join(output_dir, csv_filename)

        try:
            with open(csv_filepath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            result["output_files"].append(csv_filepath)
            result["sheet_count"] += 1
        except Exception as e:
            result["errors"].append(f"转换工作表 '{sheet_name}' 失败: {e}")

    wb.close()
    result["success"] = result["sheet_count"] > 0
    return result


# =============================================================================
# 5. CSV转Excel
# =============================================================================
def csv_to_excel(csv_path, excel_path):
    """
    将CSV文件转换为Excel文件。

    CSV文件中的第一行将被作为表头。

    Args:
        csv_path (str): 源CSV文件路径。
        excel_path (str): 输出Excel文件路径（.xlsx）。

    Returns:
        dict: 包含以下键的字典:
            - "success": 是否成功
            - "output": 输出Excel文件路径
            - "row_count": 写入的数据行数（不含表头）
            - "errors": 错误信息列表

    Example:
        >>> csv_to_excel("data.csv", "data.xlsx")
        {'success': True, 'output': 'data.xlsx', 'row_count': 100, 'errors': []}
    """
    from openpyxl import Workbook

    result = {"success": False, "output": excel_path, "row_count": 0, "errors": []}

    if not os.path.isfile(csv_path):
        result["errors"].append(f"文件不存在: {csv_path}")
        return result

    # 尝试多种编码
    rows = None
    for encoding in ["utf-8-sig", "utf-8", "gbk", "gb2312", "latin-1"]:
        try:
            with open(csv_path, "r", newline="", encoding=encoding) as f:
                reader = csv.reader(f)
                rows = list(reader)
            break
        except (UnicodeDecodeError, Exception):
            continue

    if rows is None:
        result["errors"].append("无法以常见编码读取CSV文件")
        return result

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for row in rows:
        ws.append(row)

    result["row_count"] = len(rows) - 1 if rows else 0

    try:
        wb.save(excel_path)
        result["success"] = True
    except Exception as e:
        result["errors"].append(f"保存Excel失败: {e}")

    return result


# =============================================================================
# 6. 生成会议纪要
# =============================================================================
def generate_meeting_minutes(title, attendees, agenda, decisions, action_items):
    """
    生成Markdown格式的会议纪要。

    Args:
        title (str): 会议标题。
        attendees (list): 参会人员姓名列表。
        agenda (list): 议程列表，每项为字符串。
        decisions (list): 决议列表，每项为字符串。
        action_items (list): 待办事项列表，每项为字典:
            {"task": 任务描述, "owner": 负责人, "deadline": 截止日期}。

    Returns:
        str: Markdown格式的会议纪要文本。

    Example:
        >>> minutes = generate_meeting_minutes(
        ...     "周会",
        ...     ["张三", "李四"],
        ...     ["讨论项目进度"],
        ...     ["项目按计划推进"],
        ...     [{"task": "完成模块A", "owner": "张三", "deadline": "2025-01-15"}]
        ... )
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    lines = []
    lines.append(f"# 会议纪要: {title}\n")
    lines.append(f"**日期**: {now}\n")
    lines.append(f"**参会人员**: {', '.join(attendees)}\n")
    lines.append("---\n")

    lines.append("## 议程\n")
    for i, item in enumerate(agenda, 1):
        lines.append(f"{i}. {item}")
    lines.append("")

    lines.append("## 决议\n")
    for i, decision in enumerate(decisions, 1):
        lines.append(f"{i}. {decision}")
    lines.append("")

    lines.append("## 待办事项\n")
    if action_items:
        lines.append("| # | 任务 | 负责人 | 截止日期 |")
        lines.append("|---|------|--------|----------|")
        for i, item in enumerate(action_items, 1):
            task = item.get("task", "")
            owner = item.get("owner", "")
            deadline = item.get("deadline", "")
            lines.append(f"| {i} | {task} | {owner} | {deadline} |")
    else:
        lines.append("无待办事项。")
    lines.append("")

    lines.append("---\n")
    lines.append(f"*纪要生成时间: {now}*")

    return "\n".join(lines)


# =============================================================================
# 7. 排班表生成
# =============================================================================
def create_work_schedule(employees, shifts, week_start):
    """
    生成一周排班表。

    将员工轮班分配到一周（7天）的各个班次中。

    Args:
        employees (list): 员工姓名列表。
        shifts (list): 班次列表，每项为字典 {"name": 班次名, "time": 时间段}。
            例如 [{"name": "早班", "time": "08:00-16:00"}, {"name": "晚班", "time": "16:00-24:00"}]。
        week_start (str): 一周开始日期，格式 "YYYY-MM-DD"。

    Returns:
        dict: 包含以下键的字典:
            - "schedule": 排班表，以天为键、以班次为子键的嵌套字典
            - "markdown": Markdown格式的排班表
            - "week_start": 一周开始日期

    Example:
        >>> schedule = create_work_schedule(
        ...     ["张三", "李四"],
        ...     [{"name": "早班", "time": "08:00-16:00"}],
        ...     "2025-01-06"
        ... )
    """
    start_date = datetime.strptime(week_start, "%Y-%m-%d")
    weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    schedule = {}
    emp_index = 0

    for day_offset in range(7):
        current_date = start_date + timedelta(days=day_offset)
        date_str = current_date.strftime("%Y-%m-%d")
        day_label = f"{weekdays[day_offset]}({date_str})"

        schedule[day_label] = {}
        for shift in shifts:
            # 轮班分配：每个班次轮流分配员工
            assigned_employee = employees[emp_index % len(employees)]
            schedule[day_label][shift["name"]] = {
                "employee": assigned_employee,
                "time": shift["time"],
            }
            emp_index += 1

    # 生成Markdown
    md_lines = ["# 排班表\n"]
    md_lines.append(f"**周开始日期**: {week_start}\n")

    # 表头
    header = "| 日期 |"
    separator = "|------|"
    for shift in shifts:
        header += f" {shift['name']} |"
        separator += "------|"
    md_lines.append(header)
    md_lines.append(separator)

    for day_label, day_shifts in schedule.items():
        row = f"| {day_label} |"
        for shift in shifts:
            shift_name = shift["name"]
            if shift_name in day_shifts:
                emp = day_shifts[shift_name]["employee"]
                row += f" {emp} |"
            else:
                row += " - |"
        md_lines.append(row)

    md_lines.append("")
    md_lines.append("*班次轮换自动分配*")

    return {
        "schedule": schedule,
        "markdown": "\n".join(md_lines),
        "week_start": week_start,
    }


# =============================================================================
# 8. 文件自动分类整理
# =============================================================================
def file_organizer(source_dir, categories):
    """
    根据文件扩展名将文件自动分类整理到不同目录。

    Args:
        source_dir (str): 源目录路径。
        categories (dict): 分类规则字典，键为分类名称，值为扩展名列表。
            例如 {"图片": [".jpg", ".png"], "文档": [".pdf", ".docx"]}。
            特殊键 "_default" 对应未分类文件的目录名。

    Returns:
        dict: 包含以下键的字典:
            - "total": 处理的文件总数
            - "moved": 移动的文件列表（每项为 {"file": 文件名, "category": 分类}）
            - "skipped": 跳过的文件列表
            - "errors": 错误信息列表

    Example:
        >>> result = file_organizer(
        ...     "./downloads",
        ...     {"图片": [".jpg", ".png"], "文档": [".pdf"], "_default": "其他"}
        ... )
    """
    result = {"total": 0, "moved": [], "skipped": [], "errors": []}

    if not os.path.isdir(source_dir):
        result["errors"].append(f"目录不存在: {source_dir}")
        return result

    # 构建扩展名到分类的映射
    ext_to_category = {}
    default_category = categories.get("_default", "其他")
    for category, extensions in categories.items():
        if category == "_default":
            continue
        for ext in extensions:
            ext_to_category[ext.lower()] = category

    for filename in os.listdir(source_dir):
        filepath = os.path.join(source_dir, filename)
        if not os.path.isfile(filepath):
            continue

        result["total"] += 1
        _, ext = os.path.splitext(filename)
        ext = ext.lower()

        # 确定分类
        category = ext_to_category.get(ext, default_category)

        # 创建目标目录
        dest_dir = os.path.join(source_dir, category)
        os.makedirs(dest_dir, exist_ok=True)

        dest_path = os.path.join(dest_dir, filename)

        # 处理同名文件
        if os.path.exists(dest_path):
            base, extension = os.path.splitext(filename)
            counter = 1
            while os.path.exists(dest_path):
                dest_path = os.path.join(dest_dir, f"{base}_{counter}{extension}")
                counter += 1

        try:
            shutil.move(filepath, dest_path)
            result["moved"].append({"file": filename, "category": category})
        except Exception as e:
            result["errors"].append(f"移动文件失败 {filename}: {e}")

    return result


# =============================================================================
# 9. 批量文本替换
# =============================================================================
def text_replacer(directory, find_text, replace_text, extensions=None):
    """
    批量替换指定目录下文件中的文本内容。

    Args:
        directory (str): 目标目录路径。
        find_text (str): 要查找的文本（支持正则表达式）。
        replace_text (str): 替换后的文本。
        extensions (list, optional): 要处理的文件扩展名列表（含点号）。
            例如 [".txt", ".md", ".py"]。默认为 None，处理所有文本文件。

    Returns:
        dict: 包含以下键的字典:
            - "total": 扫描的文件总数
            - "modified": 修改的文件列表（每项为 {"file": 文件路径, "replacements": 替换次数}）
            - "skipped": 跳过的文件列表
            - "errors": 错误信息列表

    Example:
        >>> result = text_replacer("./src", "old_name", "new_name", [".py"])
        {'total': 10, 'modified': [...], 'skipped': [...], 'errors': []}
    """
    result = {"total": 0, "modified": [], "skipped": [], "errors": []}

    if not os.path.isdir(directory):
        result["errors"].append(f"目录不存在: {directory}")
        return result

    # 常见文本文件扩展名
    text_extensions = extensions or [
        ".txt", ".md", ".py", ".js", ".ts", ".java", ".c", ".cpp",
        ".html", ".css", ".json", ".xml", ".yaml", ".yml", ".csv",
        ".log", ".ini", ".cfg", ".conf",
    ]

    pattern = re.compile(find_text)

    for root, dirs, files in os.walk(directory):
        for filename in files:
            _, ext = os.path.splitext(filename)
            if ext.lower() not in text_extensions:
                continue

            result["total"] += 1
            filepath = os.path.join(root, filename)

            # 尝试多种编码
            content = None
            used_encoding = None
            for encoding in ["utf-8", "gbk", "latin-1"]:
                try:
                    with open(filepath, "r", encoding=encoding) as f:
                        content = f.read()
                    used_encoding = encoding
                    break
                except (UnicodeDecodeError, Exception):
                    continue

            if content is None:
                result["skipped"].append(filepath)
                continue

            new_content, count = pattern.subn(replace_text, content)

            if count > 0:
                try:
                    with open(filepath, "w", encoding=used_encoding) as f:
                        f.write(new_content)
                    result["modified"].append({"file": filepath, "replacements": count})
                except Exception as e:
                    result["errors"].append(f"写入文件失败 {filepath}: {e}")
            else:
                result["skipped"].append(filepath)

    return result


# =============================================================================
# 10. 密码生成器
# =============================================================================
def password_generator(length=12, use_upper=True, use_lower=True,
                       use_digits=True, use_symbols=True):
    """
    生成随机密码。

    Args:
        length (int): 密码长度，默认12。
        use_upper (bool): 是否包含大写字母，默认True。
        use_lower (bool): 是否包含小写字母，默认True。
        use_digits (bool): 是否包含数字，默认True。
        use_symbols (bool): 是否包含特殊符号，默认True。

    Returns:
        dict: 包含以下键的字典:
            - "password": 生成的密码字符串
            - "length": 密码长度
            - "strength": 密码强度评级（"弱"/"中"/"强"/"极强"）
            - "character_sets": 使用的字符集列表

    Example:
        >>> result = password_generator(length=16)
        {'password': 'aB3$kL9@mN2#pQ7x', 'length': 16, 'strength': '极强', ...}
    """
    if length < 4:
        length = 4

    char_sets = []
    used_sets = []

    if use_upper:
        char_sets.append(string.ascii_uppercase)
        used_sets.append("大写字母")
    if use_lower:
        char_sets.append(string.ascii_lowercase)
        used_sets.append("小写字母")
    if use_digits:
        char_sets.append(string.digits)
        used_sets.append("数字")
    if use_symbols:
        char_sets.append("!@#$%^&*()-_=+[]{}|;:,.<>?")
        used_sets.append("特殊符号")

    if not char_sets:
        char_sets.append(string.ascii_lowercase)
        used_sets.append("小写字母(默认)")

    all_chars = "".join(char_sets)

    # 确保每个字符集至少出现一次
    password_chars = []
    for char_set in char_sets:
        password_chars.append(random.choice(char_set))

    # 填充剩余长度
    remaining = length - len(password_chars)
    for _ in range(remaining):
        password_chars.append(random.choice(all_chars))

    # 打乱顺序
    random.shuffle(password_chars)
    password = "".join(password_chars)

    # 评估强度
    pool_size = len(all_chars)
    import math
    entropy = length * math.log2(pool_size) if pool_size > 0 else 0

    if entropy < 40:
        strength = "弱"
    elif entropy < 60:
        strength = "中"
    elif entropy < 80:
        strength = "强"
    else:
        strength = "极强"

    return {
        "password": password,
        "length": length,
        "strength": strength,
        "character_sets": used_sets,
    }


# =============================================================================
# 主入口
# =============================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("办公效率工具箱 (office-helper)")
    print("=" * 60)
    print("可用工具:")
    tools = [
        "1. batch_rename_files      - 批量重命名文件",
        "2. merge_pdf_files         - 合并PDF文件",
        "3. split_pdf_file          - 拆分PDF文件",
        "4. excel_to_csv            - Excel转CSV",
        "5. csv_to_excel            - CSV转Excel",
        "6. generate_meeting_minutes- 生成会议纪要",
        "7. create_work_schedule    - 排班表生成",
        "8. file_organizer          - 文件自动分类整理",
        "9. text_replacer           - 批量文本替换",
        "10. password_generator     - 密码生成器",
    ]
    for tool in tools:
        print(f"  {tool}")
    print("=" * 60)

    # 演示：生成密码
    print("\n演示 - 生成密码:")
    pwd_result = password_generator(length=16)
    print(f"  密码: {pwd_result['password']}")
    print(f"  强度: {pwd_result['strength']}")
    print(f"  字符集: {', '.join(pwd_result['character_sets'])}")
